import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import permissions


class QueryTokenMixin:
    """Allow authentication via ?token= query parameter for browser tab downloads."""
    
    def initialize_request(self, request, *args, **kwargs):
        drf_request = super().initialize_request(request, *args, **kwargs)
        token = request.GET.get('token')
        if token and not drf_request.auth:
            try:
                from jose import jwt as jose_jwt
                from django.conf import settings
                payload = jose_jwt.decode(
                    token, settings.JWT_PUBLIC_KEY, algorithms=[settings.JWT_ALGORITHM]
                )
                from django.contrib.auth import get_user_model
                User = get_user_model()
                user = User.objects.get(id=payload['user_id'])
                drf_request.user = user
                drf_request._force_auth_user = user
                drf_request.auth = token
            except Exception:
                pass
        return drf_request


def style_header(ws, headers, row=1):
    """Apply professional styling to header row."""
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max(len(str(header)) + 4, 14)


def make_excel_response(wb, filename):
    """Convert workbook to HTTP response."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class ExportBillingDataView(QueryTokenMixin, APIView):
    """Export all billing data as Excel spreadsheet."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user_role = request.user.role.name.upper() if request.user.role else ''
        if not request.user.is_staff and user_role not in ['ADMIN']:
            return HttpResponse('Unauthorized', status=403)

        from apps.billing.models import Bill
        bills = Bill.objects.all().select_related('customer__user').order_by('-created_at')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Bills'

        # Title row
        ws.merge_cells('A1:H1')
        title_cell = ws.cell(row=1, column=1, value=f'AquaBill AI — Billing Report ({datetime.now().strftime("%Y-%m-%d")})')
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='1E3A8A')
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Bill ID', 'Customer', 'Email', 'Consumption (m³)', 'Amount (ETB)', 'Status', 'Due Date', 'Created']
        style_header(ws, headers, row=3)

        for idx, bill in enumerate(bills, 4):
            customer_name = ''
            email = ''
            if bill.customer and hasattr(bill.customer, 'user') and bill.customer.user:
                customer_name = f"{bill.customer.user.first_name} {bill.customer.user.last_name}".strip()
                email = bill.customer.user.email

            ws.cell(row=idx, column=1, value=str(bill.id)[:8])
            ws.cell(row=idx, column=2, value=customer_name or 'N/A')
            ws.cell(row=idx, column=3, value=email)
            ws.cell(row=idx, column=4, value=float(bill.consumption))
            ws.cell(row=idx, column=5, value=float(bill.total_amount))
            status_cell = ws.cell(row=idx, column=6, value=bill.status)
            ws.cell(row=idx, column=7, value=str(bill.due_date) if bill.due_date else 'N/A')
            ws.cell(row=idx, column=8, value=bill.created_at.strftime('%Y-%m-%d %H:%M'))

            # Color-code status
            if bill.status == 'PAID':
                status_cell.font = Font(color='059669', bold=True)
            elif bill.status == 'OVERDUE':
                status_cell.font = Font(color='DC2626', bold=True)
            elif bill.status == 'UNPAID':
                status_cell.font = Font(color='D97706', bold=True)

        filename = f'AquaBill_Bills_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return make_excel_response(wb, filename)


class ExportCustomerListView(QueryTokenMixin, APIView):
    """Export customer list as Excel spreadsheet."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user_role = request.user.role.name.upper() if request.user.role else ''
        if not request.user.is_staff and user_role not in ['ADMIN']:
            return HttpResponse('Unauthorized', status=403)

        from apps.accounts.models import Customer
        customers = Customer.objects.all().select_related('user')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Customers'

        ws.merge_cells('A1:G1')
        title_cell = ws.cell(row=1, column=1, value=f'AquaBill AI — Customer List ({datetime.now().strftime("%Y-%m-%d")})')
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='1E3A8A')
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Name', 'Email', 'Phone', 'Class', 'Active', 'Joined', 'Meters']
        style_header(ws, headers, row=3)

        from apps.metering.models import Meter
        for idx, cust in enumerate(customers, 4):
            name = f"{cust.user.first_name} {cust.user.last_name}".strip() if cust.user else 'N/A'
            meter_count = Meter.objects.filter(customer=cust).count()

            ws.cell(row=idx, column=1, value=name)
            ws.cell(row=idx, column=2, value=cust.user.email if cust.user else 'N/A')
            ws.cell(row=idx, column=3, value=cust.phone or 'N/A')
            ws.cell(row=idx, column=4, value=cust.customer_class or 'RESIDENT')
            ws.cell(row=idx, column=5, value='Yes' if cust.user and cust.user.is_active else 'No')
            ws.cell(row=idx, column=6, value=cust.user.date_joined.strftime('%Y-%m-%d') if cust.user else 'N/A')
            ws.cell(row=idx, column=7, value=meter_count)

        filename = f'AquaBill_Customers_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return make_excel_response(wb, filename)


class ExportAnomalyReportView(QueryTokenMixin, APIView):
    """Export anomaly/alert reports as Excel spreadsheet."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user_role = request.user.role.name.upper() if request.user.role else ''
        if not request.user.is_staff and user_role not in ['ADMIN']:
            return HttpResponse('Unauthorized', status=403)

        from apps.billing.models import WaterAlert
        alerts = WaterAlert.objects.all().select_related('customer__user').order_by('-created_at')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Anomaly Reports'

        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1, value=f'AquaBill AI — Anomaly Reports ({datetime.now().strftime("%Y-%m-%d")})')
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='1E3A8A')
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Alert ID', 'Customer', 'Type', 'Message', 'Resolved', 'Date']
        style_header(ws, headers, row=3)

        for idx, alert in enumerate(alerts, 4):
            customer_name = ''
            if alert.customer and hasattr(alert.customer, 'user') and alert.customer.user:
                customer_name = f"{alert.customer.user.first_name} {alert.customer.user.last_name}".strip()

            ws.cell(row=idx, column=1, value=str(alert.id)[:8])
            ws.cell(row=idx, column=2, value=customer_name or 'N/A')
            ws.cell(row=idx, column=3, value=alert.alert_type)
            ws.cell(row=idx, column=4, value=alert.message[:100] if alert.message else '')
            resolved_cell = ws.cell(row=idx, column=5, value='Yes' if alert.is_resolved else 'No')
            ws.cell(row=idx, column=6, value=alert.created_at.strftime('%Y-%m-%d %H:%M'))

            if not alert.is_resolved:
                resolved_cell.font = Font(color='DC2626', bold=True)

        filename = f'AquaBill_Anomalies_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return make_excel_response(wb, filename)
